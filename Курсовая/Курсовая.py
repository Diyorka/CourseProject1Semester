import os
import openpyxl as xl
from prettytable import PrettyTable

print('Hello! Welcome to "StayFit"!')

def director():
    print('Hello, dear Director!')
    while True:
        w=xl.load_workbook('loginpass.xlsx')
        log = input('Enter your login:\n')
        sh = w['director']
        for i in range(2, sh.max_row + 1):
            if log == sh.cell(row=i, column=1).value:
                logi = i
                break
            else:
                logi = 'No'
        if logi != 'No':
            while True:
                pas = input('Enter your password:\n')
                if pas == sh.cell(row=logi, column=2).value:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print("Welcome, dear Director!")
                    print("Please, enter the number of menu from 1 to 7", "1 - List of managers", "2 - List of s. assistants", "3 - List of clients", "4 - List of procedures", "5 - Info about a client", "6 - Info about a s. assistant", "7 - Leave the menu", sep ='\n')
                    while True:
                        mnum = int(input("Please, enter the number of menu from 1 to 6, if you finish enter 7:\n"))
                        if mnum == 1:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('1. List of managers:')
                            wb = xl.load_workbook('loginpass.xlsx')
                            sh = wb['manager']
                            r = 2
                            if sh.cell(row=r, column=1).value==None:
                                table = PrettyTable()
                                table.field_names = ['Managers:']
                                table.add_row(['There is not any manager'])
                            else:
                                table = PrettyTable()
                                table.field_names = ['Managers:']
                                while sh.cell(row=r, column=1).value != None:
                                    table.add_row([sh.cell(row=r,column=1).value])
                                    r+=1
                            print(table)
                            print()
                        elif mnum == 2:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('2. List of s. assistants:')
                            wb = xl.load_workbook('loginpass.xlsx')
                            sh = wb['sassistant']
                            r = 2
                            if sh.cell(row=r, column=1).value==None:
                                table = PrettyTable()
                                table.field_names = ['Sanatorium assistants:']
                                table.add_row(['There is not any sanatorium assistant'])
                            else:
                                table = PrettyTable()
                                table.field_names = ['Sanatorium assistants:']
                                while sh.cell(row=r, column=1).value != None:
                                    table.add_row([sh.cell(row=r, column=1).value])
                                    r+=1
                            print(table)
                            print()
                        elif mnum == 3:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('3. List of clients:')
                            wb = xl.load_workbook('loginpass.xlsx')
                            sh = wb['client']
                            r = 2
                            if sh.cell(row=r, column=1).value==None:
                                table=PrettyTable()
                                table.field_names = ['Clients:']
                                table.add_row(['There is not any client'])
                            else:
                                table = PrettyTable()
                                table.field_names = ['Clients:']
                                while sh.cell(row=r, column=1).value != None:
                                    table.add_row([sh.cell(row=r, column=1).value])
                                    r+=1
                            print(table)
                            print()
                        elif mnum == 4:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('4. List of procedures:')
                            wb = xl.load_workbook('price_list.xlsx')
                            sh = wb['price']
                            sp = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
                            table = PrettyTable(sp)
                            for a, b in sh['A2': 'B6']:
                                table.add_row(
                                    [a.value, b.value])
                            print(table)
                            print()
                        elif mnum == 5:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('5. Looking for a client:')
                            while True:
                                lfc = input('Enter name of the client:\n').capitalize()
                                wb = xl.load_workbook('loginpass.xlsx')
                                sh = wb['client']
                                r = 2
                                while sh.cell(row=r, column=1).value != None:
                                    if lfc == sh.cell(row=r, column=1).value:
                                        exist = r-2
                                        break
                                    else:
                                        exist = 'No'
                                    r+=1
                                if exist != 'No':
                                    wb = xl.load_workbook('client_info.xlsx')
                                    sh = wb[lfc]
                                    r = 2
                                    sp = [sh.cell(row=1, column=1).value]
                                    table = PrettyTable(sp)
                                    while sh.cell(row=r,column=1).value !=None:
                                        table.add_row([sh.cell(row=r,column=1).value])
                                        r+=1
                                    print(table)
                                    print()
                                    break
                                else:
                                    os.system('cls' if os.name == 'nt' else 'clear')
                                    print('Please, write correct name')
                        elif mnum == 6:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('6. Looking for a s. assistant:')
                            while True:
                                lfs = input('Enter name of the s. assistant:\n').capitalize()
                                wb = xl.load_workbook('loginpass.xlsx')
                                sh = wb['sassistant']
                                r = 2
                                while sh.cell(row=r, column=1).value != None:
                                    if lfs == sh.cell(row=r, column=1).value:
                                        exist = r-2
                                        break
                                    else:
                                        exist = 'No'
                                    r+=1
                                if exist != 'No':
                                    wb = xl.load_workbook('sassistant_info.xlsx')
                                    sh = wb[lfs]
                                    r = 2
                                    sp = [sh.cell(row=1, column=1).value]
                                    table = PrettyTable(sp)
                                    while sh.cell(row=r, column=1).value != None:
                                        table.add_row([sh.cell(row=r, column=1).value])
                                        r += 1
                                    print(table)
                                    table2=PrettyTable()
                                    table2.field_names=['Salary']
                                    table2.add_row([sh.cell(row=2, column=2).value])
                                    print(table2)
                                    print()
                                    break
                                else:
                                    os.system('cls' if os.name == 'nt' else 'clear')
                                    print('Please, write correct name')
                        elif mnum == 7:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('You left your account.')
                            main()
                            break
                        else:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('Enter number from 1 to 6')
                    break
                else:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print('Write correct password')
            break
        else:
            os.system('cls' if os.name == 'nt' else 'clear')
            print('Write correct login')


def client():  # Аккаунт клиента
    print('Hello, dear visitor!')
    while True:
        w = xl.load_workbook('loginpass.xlsx')
        log = input("Enter your login:\n") # Проверка существования логина
        sh = w['client']
        for i in range(2, sh.max_row + 1):
            if log == sh.cell(row=i, column=1).value:
                logi = i
                break
            else:
                logi = 'No'
        if logi != 'No':
            while True:  # Проверка существования пароля
                pas = input("Enter your password:\n")
                if pas == sh.cell(row=logi, column=2).value:  # Проверка индекса логина и пароля в списке:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print("Welcome, dear Visitor!")
                    print("Please, enter the number of menu from 1 to 7", "1 - Visiting history",
                          "2 - The last visiting", "3 - History of purchases", "4 - Your timetable", "5 - Info about you",
                           "6 - Buying a procedure", "7 - Info about our sanatorium assistants", "8 - Leaving the menu", sep='\n')
                    while True:
                        mnum = int(input("Please, enter the number of menu from 1 to 7, if you finish enter 8:\n"))
                        if mnum == 1:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('1. Visiting history: ')
                            wb = xl.load_workbook('client_info.xlsx')
                            sh = wb[f'{log}']
                            r = 2
                            if sh.cell(row=r, column=3).value == None:
                                print('Your history is empty!')
                            else:
                                table = PrettyTable()
                                table.field_names = ['History']
                                while sh.cell(row=r, column=3).value != None:
                                    table.add_row([sh.cell(row=r, column=3).value])
                                    r += 1
                                r-=1
                                print(table)
                            print()
                        elif mnum == 2:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('2. The last visiting:')
                            wb = xl.load_workbook('client_info.xlsx')
                            sh = wb[f'{log}']
                            if sh.cell(row=r, column=3).value == None:
                                print('Your history is empty!')
                            else:
                                print(sh.cell(row=r, column=3).value)
                            print()
                        elif mnum == 3:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('3. Purchased procedures: ')
                            wb = xl.load_workbook('client_info.xlsx')
                            sh = wb[f'{log}']
                            r = 2
                            if sh.cell(row=2, column=4).value == None:
                                print('History is empty')
                            else:
                                table = PrettyTable()
                                table.field_names = ['Purchased procedures:']
                                while sh.cell(row=r, column=4).value != None:
                                    table.add_row([sh.cell(row=r, column=4).value])
                                    r += 1
                            print(table)
                            print()
                        elif mnum == 4:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('4. Timetable:')
                            wb = xl.load_workbook('client_info.xlsx')
                            sh = wb[f'{log}']
                            r = 2
                            import random
                            if sh.cell(row=r, column=2).value == None:
                                print('Your timetable is empty!')
                                print()
                                while True:
                                    choosing = input('Do you want to choose timetable? (enter "yes" or "no")\n').lower()
                                    if choosing == 'yes':
                                        visittimes = int(input('How many times a week do you want to visit our fitness center?\n'))
                                        days = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday')
                                        for _ in range(visittimes):
                                            while True:
                                                day = input(f'Choose the {r-1} day(example: {random.choice(days)}):\n').capitalize()
                                                if day in days:
                                                    break
                                                elif day == 'Sunday':
                                                    print('We do not work on Sunday, choose another day.')
                                                else:
                                                    print('Write correctly!')

                                            sh.cell(row=r, column=2).value = day + ': ' + input('Choose time(example: 18:00-20:00):\n')
                                            r+=1
                                        wb.save('client_info.xlsx')
                                        print('Your timetable was successfully saved!')
                                        break
                                    elif choosing == 'no':
                                        break
                                    else:
                                        print('Write correctly!')
                                r = 2
                            else:
                                table = PrettyTable()
                                table.field_names = ['Your timetable:']
                                while sh.cell(row=r, column=2).value != None:
                                    table.add_row([sh.cell(row=r, column=2).value])
                                    r += 1
                                r=2
                                print(table)
                                print()
                                while True:
                                    changing = input('Do you want to change your timetable?(enter "yes" or "no")\n').lower()
                                    if changing == 'yes':
                                        while sh.cell(row=r, column=2).value != None:
                                            sh.cell(row=r, column=2).value = None
                                            r+=1
                                        visisttimes = int(input('How many times a week do you want to visit our fitness center?\n'))
                                        days = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday')
                                        r=2
                                        for _ in range(visisttimes):
                                            while True:
                                                day = input(f'Choose the {r-1} day(example: {random.choice(days)}):\n').capitalize()
                                                if day in days:
                                                    break
                                                elif day == 'Sunday':
                                                    print('We do not work on Sunday, choose another day.')
                                                else:
                                                    print('Write correctly!')

                                            sh.cell(row=r, column=2).value = day + ': ' + input('Choose time(example: 18:00-20:00):\n')
                                            r += 1
                                        wb.save('client_info.xlsx')
                                        print('Successfully changed!')
                                        break
                                    elif changing == 'no':
                                        break
                                    else:
                                        print('Write correctly!')
                            print()
                        elif mnum == 5:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('5. Info about you:')
                            wb = xl.load_workbook('client_info.xlsx')
                            sh = wb[f'{log}']
                            r = 2
                            table = PrettyTable()
                            table.field_names = ['Your info:']
                            while sh.cell(row=r, column=1).value != None:
                                table.add_row([sh.cell(row=r, column=1).value])
                                r += 1
                            print(table)
                            print()
                        elif mnum == 6:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('6. Buying a procedure:')
                            while True:
                                w = xl.load_workbook('price_list.xlsx')
                                sheet = w['price']
                                sp = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
                                table = PrettyTable(sp)
                                for a, b in sheet['A2': 'B6']:
                                    table.add_row([a.value, b.value])
                                print(table)
                                print()
                                procname = input('Enter name of the procedure or enter "leave" to leave this menu:\n').capitalize()
                                if procname != 'Leave':
                                    r = 2
                                    checkproc = 'NO'
                                    while sheet.cell(row=r, column=1).value != None:
                                        if sheet.cell(row=r, column=1).value == procname:
                                            checkproc = 'YES'
                                            break
                                        r += 1
                                    if checkproc == 'YES':
                                        w2 = xl.load_workbook('client_info.xlsx')
                                        sh2 = w2[f'{log}']
                                        r = 2
                                        while sh2.cell(row=r, column=4).value != None:
                                            r += 1
                                        w = xl.load_workbook('price_list.xlsx')
                                        wsh = w['time']
                                        sp = [wsh.cell(row=1, column=i).value for i in range(1, wsh.max_column + 1)]
                                        table = PrettyTable(sp)
                                        for a, b, c, d, e, f, g in wsh['A2': 'G6']:
                                            table.add_row(
                                                [a.value, b.value, c.value, d.value, e.value, f.value, g.value])
                                        print(table)
                                        print()
                                        while True:
                                            day = input('Choose the day:\n').capitalize()
                                            days = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday')
                                            if day in days:
                                                ptime = input('Choose procedure time(example: 14:00):\n')
                                                sh2.cell(row=r, column=4).value = procname + f', {day}, {ptime}'
                                                print('Successfully bought!')
                                                print()
                                                w2.save('client_info.xlsx')
                                                break
                                            else:
                                                print('Write correctly!')
                                        break
                                    else:
                                        print('Enter correct name of the procedure!')
                                        print()
                                else:
                                    print()
                                    break
                        elif mnum == 7:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('7. Info about our sanatorium assistants:')
                            w=xl.load_workbook('sassistant_info.xlsx')
                            table = PrettyTable()
                            table.field_names = ['Name', 'Phone number']
                            for i in range(len(w.sheetnames)):
                                w.active=i
                                sh=w.active
                                table.add_row([sh.cell(row=2, column=1).value,sh.cell(row=4,column=1).value])
                            print(table)
                            print()
                        elif mnum == 8:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('You left your account.')
                            main()
                            break
                        else:
                            print("Enter number from 1 to 7")
                    break
                else:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print('Write correct password\n')
            break
        else:
            os.system('cls' if os.name == 'nt' else 'clear')
            print('Write correct login\n')


def manager():  # Аккаунт менеджера
    print('Hello, dear manager!')
    while True:  # Проверка существования логина
        w = xl.load_workbook('loginpass.xlsx')
        log = input("Enter your login:\n")
        sh = w['manager']
        for i in range(2, sh.max_row + 1):
            if log == sh.cell(row=i, column=1).value:
                logi = i
                break
            else:
                logi = 'No'
        if logi != 'No':
            while True:  # Проверка существования пароля
                pas = input("Enter your password:\n")
                if pas == sh.cell(row=logi, column=2).value:  # Проверка индекса логина и пароля в списке
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print(f"Welcome, dear Manager!")
                    print("Please, enter the number of menu from 1 to 8", "1 - Visitors list", "2 - Count of visitors",
                          "3 - Visitor searching", "4 - Price changing", "5 - Procedure time changing",
                          "6 - The visitor with the most visited days", "7 - The visitor with the least visited days",
                          "8 - Add or delete s. assistant", "9 - S. assistant's salary changing","10 - Leave the menu", sep='\n')
                    while True:
                        mnum = int(input("Please, enter the number of menu from 1 to 9, if you finish enter 10:\n"))
                        if mnum == 1:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('1. Visitors list:')
                            sh = w['client']
                            table = PrettyTable()
                            table.field_names = ['Clients:']
                            r=2
                            while sh.cell(row=r, column=1).value != None:
                                table.add_row([sh.cell(row=r, column=1).value])
                                r+=1
                            print(table)
                            print()
                        elif mnum == 2:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('2. Count of visitors:')
                            sh = w['client']
                            print(sh.max_row - 1)
                            print()
                        elif mnum == 3:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('3. Visitor searching: ')
                            while True:
                                sh = w['client']
                                cname = input('Please, enter name of the visitor or enter "leave" to leave this menu:\n')
                                if cname != 'leave':
                                    for i in range(2, sh.max_row + 1):
                                        if cname == sh.cell(row=i, column=1).value:
                                            cnamei = i
                                            break
                                        else:
                                            cnamei = 'No'
                                    if cnamei != 'No':
                                        os.system('cls' if os.name == 'nt' else 'clear')
                                        wb = xl.load_workbook('client_info.xlsx')
                                        print('Info about the visitor:')
                                        csh = wb[f'{cname}']
                                        r=2
                                        table = PrettyTable()
                                        table.field_names = ["Visitor's info:"]
                                        while csh.cell(row=r, column=1).value!=None:
                                            table.add_row([csh.cell(row=r,column=1).value])
                                            r+=1
                                        print(table)
                                        print()
                                        break
                                    else:
                                        os.system('cls' if os.name == 'nt' else 'clear')
                                        print('Write correct name\n')
                                else:
                                    print()
                                    break
                        elif mnum == 4:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('4. Price changing\n')
                            w = xl.load_workbook('price_list.xlsx')
                            sh = w['price']
                            sp = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
                            table = PrettyTable(sp)
                            for a, b in sh['A2': 'B6']:
                                table.add_row(
                                    [a.value, b.value])
                            print(table)
                            print()
                            while True:
                                namep = input('Enter name of the procedure, which price you want to change or enter "leave" to leave this menu:\n').lower()
                                if namep != 'leave':
                                    for i in range(2, sh.max_row + 1):
                                        if namep.capitalize() == sh.cell(row=i, column=1).value:
                                            namep1 = i
                                            break
                                        else:
                                            namep1 = 'No'
                                    if namep1 != 'No':
                                        nprice = input('Enter new price:\n')
                                        if nprice.isdigit():
                                            sh.cell(row=namep1, column=2).value = (nprice+'$')
                                            w.save('price_list.xlsx')
                                            break
                                        else:
                                            print('Please write correct price')
                                            print()
                                    else:
                                        print('Please write correct name')
                                        print()
                                else:
                                    print()
                                    break
                        elif mnum == 5:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('5. Procedure time changing\n')
                            w = xl.load_workbook('price_list.xlsx')
                            sh = w['time']
                            sp = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
                            table = PrettyTable(sp)
                            for a, b, c, d, e, f, g in sh['A2': 'G6']:
                                table.add_row(
                                    [a.value, b.value, c.value, d.value, e.value, f.value, g.value])
                            print(table)
                            print()
                            while True:
                                w = xl.load_workbook('price_list.xlsx')
                                sh = w['time']
                                t = input('Enter name of the procedure, which time you want to change or enter "leave" to leave this menu:\n').lower()
                                if t != 'leave':
                                    for i in range(2, sh.max_row + 1):
                                        if t.capitalize() == sh.cell(row=i, column=1).value:
                                            ti = i
                                            break
                                        else:
                                            ti = 'No'
                                    if ti != 'No':
                                        while True:
                                            d = input('Enter the day of week:\n')
                                            for i in range(2, sh.max_column + 1):
                                                if d.capitalize() == sh.cell(row=1, column=i).value:
                                                    dayi = i
                                                    break
                                                else:
                                                    dayi = 'No'
                                            if dayi != 'No':
                                                ntime = input('Enter new time(Example: 10:00-20:00):\n')
                                                sh.cell(row=ti, column=dayi).value = ntime
                                                w.save('price_list.xlsx')
                                                print()
                                                break
                                            else:
                                                print('Write correct day!\n')
                                        break
                                    else:
                                        print('Write correct name!\n')
                                else:
                                    print()
                                    break
                        elif mnum == 6:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('6. The visitor with the most visited days: ')
                            wb = xl.load_workbook('client_info.xlsx')
                            m=0
                            ni=0
                            for i in range(len(wb.sheetnames)):
                                wb.active=i
                                sh = wb.active
                                r=2
                                while sh.cell(row=r, column=3).value != None:
                                    r+=1
                                if (r-2)>m:
                                    m=(r-2)
                                    ni=i
                            wb.active = ni
                            sh = wb.active
                            print(sh.cell(row=2, column=1).value)
                            print()
                        elif mnum == 7:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('7. The visitor with the least visited days: ')
                            wb = xl.load_workbook('client_info.xlsx')
                            m=999999999
                            ni=0
                            for i in range(len(wb.sheetnames)):
                                wb.active=i
                                sh = wb.active
                                r=2
                                while sh.cell(row=r, column=3).value != None:
                                    r+=1
                                if (r-2)<m:
                                    m=(r-2)
                                    ni=i
                            wb.active = ni
                            sh = wb.active
                            print(sh.cell(row=2, column=1).value)
                            print()
                        elif mnum == 8:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            while True:
                                addordel = input('8. Do you want to add or delete s. assistant?(Enter "add" or "del" or "leave" if you want to leave this menu):\n').lower()
                                if addordel == 'add':
                                    while True:
                                        addname = input('Enter name of the new s. assistant:\n').capitalize()
                                        addpass = input('Enter password of the new s. assistant:\n')
                                        addfname = input('Enter full name of the new s. assistant(example: Umurzakov Diyor):\n').title()
                                        w=xl.load_workbook('loginpass.xlsx')
                                        sh=w['sassistant']
                                        r = 2
                                        while sh.cell(row=r, column=1).value != None:
                                            r+=1
                                        sh.cell(row=r, column=1).value = addname
                                        sh.cell(row=r, column=2).value = addpass
                                        w.save('loginpass.xlsx')
                                        w=xl.load_workbook('sassistant_info.xlsx')
                                        w.create_sheet(addname)
                                        sh=w[addname]
                                        sh.cell(row=1, column=1).value = 'Info'
                                        sh.cell(row=1, column=2).value = 'Salary'
                                        sh.cell(row=2,column=1).value = addfname
                                        sh.cell(row=3, column=1).value = 'Birthday: ' + input('Enter his(her) birthday(example: 01.01.2000):\n')
                                        sh.cell(row=4, column=1).value = 'Phone num.: ' + input('Enter his(her) phone number(without +996 and 0):\n')
                                        sh.cell(row=2,column=2).value = int(input('Enter his(her) salary(example: 20000):\n'))
                                        w.save('sassistant_info.xlsx')
                                        print('Account successfully created!\n')
                                        break
                                    break
                                elif addordel == 'del':
                                    while True:
                                        delname = input('Enter name of the s. assistant to delete:\n')
                                        w=xl.load_workbook('loginpass.xlsx')
                                        sure = input('Are you sure?(Enter "yes" or "no")\n').lower()
                                        if sure == 'yes':
                                            sh=w['sassistant']
                                            r = 2
                                            while sh.cell(row=r, column=1).value != None:
                                                if sh.cell(row=r, column=1).value == delname:
                                                    res = 'Yes'
                                                    break
                                                else:
                                                    res = 'No'
                                                r+=1
                                            if res == 'Yes':
                                                sh.delete_rows(r)
                                                w.save('loginpass.xlsx')
                                                w=xl.load_workbook('sassistant_info.xlsx')
                                                del w[delname]
                                                w.save('sassistant_info.xlsx')
                                                print('Successfully deleted!')
                                                print()
                                                break
                                            else:
                                                print('Write correct name!')
                                        else:
                                            print('S. assistant will not be deleted.')
                                            print()
                                            break
                                    break
                                elif addordel == 'leave':
                                    break
                                else:
                                    os.system('cls' if os.name == 'nt' else 'clear')
                                    print('Write correctly!')
                        elif mnum == 9:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print("9. S. assistant's salary changing:")
                            w = xl.load_workbook('loginpass.xlsx')
                            sh=w['sassistant']
                            while True:
                                name = input('Enter name of the s. assistant:\n')
                                r=2
                                while sh.cell(row=r, column=1).value != None:
                                    if sh.cell(row=r, column=1).value == name:
                                        exist = 'Yes'
                                    else:
                                        exist = 'No'
                                    r+=1
                                if exist == 'Yes':
                                    w=xl.load_workbook('sassistant_info.xlsx')
                                    sh=w[name]
                                    table = PrettyTable()
                                    table.field_names=['Info', 'Salary']
                                    table.add_row([sh.cell(row=2, column=1).value, sh.cell(row=2, column=2).value])
                                    print(table)
                                    nsalary = int(input('Enter new salary(example: 25000):\n'))
                                    sh.cell(row=2, column=2).value = nsalary
                                    print('Successfully changed!\n')
                                    w.save('sassistant_info.xlsx')
                                    break
                                else:
                                    print('Enter correct name!\n')
                        elif mnum == 10:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('You left your account.')
                            main()
                            break
                        else:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print("Enter number from 1 to 9")
                    break
                else:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print('Write correct password!\n')
            break
        else:
            os.system('cls' if os.name == 'nt' else 'clear')
            print('Write correct login!\n')


def sassistant():  # Аккаунт ассистента
    print('Hello, dear sanatorium assistant!')
    while True:  # Проверка существования логина
        w = xl.load_workbook('loginpass.xlsx')
        log = input("Enter your login:\n")
        sh = w['sassistant']
        for i in range(2, sh.max_row + 1):
            if log == sh.cell(row=i, column=1).value:
                logi = i
                break
            else:
                logi = 'No'
        if logi != 'No':
            while True:  # Проверка существования пароля
                pas = input("Enter your password:\n")
                if pas == sh.cell(row=logi, column=2).value:  # Проверка индекса логина и пароля в списке
                    os.system('cls' if os.name == 'nt' else 'clear')
                    print(f"Welcome, dear Sanatorium Assistant!")
                    print("Please, enter the number of menu from 1 to 7", "1 - Purchased procedures list",
                          "2 - Looking for client's timetable", "3 - Show all procedures", "4 - Timetable of procedures",
                          "5 - Buying a procedure", "6 - Looking for a procedure", "7 - Info about you", "8 - Leaving the menu", sep='\n')
                    while True:
                        mnum = int(input("Please, enter the number of menu from 1 to 7, if you finish enter 8:\n"))
                        if mnum == 1:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('1. Purchased procedures list:\n')
                            wb = xl.load_workbook('client_info.xlsx')
                            count = len(wb.sheetnames)
                            for i in range(count):
                                wb.active=i
                                wbsh=wb.active
                                r=2
                                if wbsh.cell(row=r, column=4).value != None:
                                    table = PrettyTable()
                                    table.field_names = [f"{wbsh.cell(row=2, column=1).value}'s purchased procedures:"]
                                    r=2
                                    while wbsh.cell(row=r, column=4).value != None:
                                        table.add_row([wbsh.cell(row=r, column=4).value])
                                        r+=1
                                    print(table)
                                    print()
                                else:
                                    table = PrettyTable()
                                    table.field_names = [f"{wbsh.cell(row=2, column=1).value}'s purchased procedures:"]
                                    table.add_row(['Empty'])
                                    print(table)
                                    print()
                        elif mnum == 2:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print("2. Looking for client's timetable:")
                            while True:
                                lfc = input('Enter name of the client or enter "leave" to leave this menu:\n').capitalize()
                                print()
                                if lfc != 'Leave':
                                    wb = xl.load_workbook('loginpass.xlsx')
                                    sh = wb['client']
                                    r = 2
                                    while sh.cell(row=r, column=1).value != None:
                                        if lfc == sh.cell(row=r, column=1).value:
                                            exist = r-2
                                            break
                                        else:
                                            exist = 'No'
                                        r+=1
                                    if exist != 'No':
                                        wb = xl.load_workbook('client_info.xlsx')
                                        sh = wb[lfc]
                                        r=2
                                        if sh.cell(row=2, column=2).value!=None:
                                            table = PrettyTable()
                                            table.field_names = [f"{sh.cell(row=2, column=1).value}'s timetable"]
                                            while sh.cell(row=r, column=2).value != None:
                                                table.add_row([sh.cell(row=r, column=2).value])
                                                r +=1
                                            print(table)
                                            print()
                                            break
                                        else:
                                            table = PrettyTable()
                                            table.field_names = [f"{sh.cell(row=2, column=1).value}'s timetable"]
                                            table.add_row(['Timetable is empty.'])
                                            print(table)
                                        print()
                                    else:
                                        os.system('cls' if os.name == 'nt' else 'clear')
                                        print('Write correct name!')
                                else:
                                    print()
                                    break
                        elif mnum == 3:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('3. All procedures:')
                            wb = xl.load_workbook('price_list.xlsx')
                            sh = wb['price']
                            r = 2
                            table = PrettyTable()
                            table.field_names = ['Procedures list:']
                            while sh.cell(row=r, column=1).value != None:
                                table.add_row([sh.cell(row=r, column=1).value])
                                r += 1
                            print(table)
                            print()
                        elif mnum == 4:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('4. Timetable of procedures:')
                            w = xl.load_workbook('price_list.xlsx')
                            sh = w['time']
                            sp = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
                            table = PrettyTable(sp)
                            for a, b, c, d, e, f, g in sh['A2': 'G6']:
                                table.add_row([a.value, b.value, c.value, d.value, e.value, f.value, g.value])
                            print(table)
                            print()
                        elif mnum == 5:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('5. Buying a procedure:')
                            while True:
                                wb = xl.load_workbook('loginpass.xlsx')
                                name = input('Enter name of the client, who buys a procedure or enter "leave" to leave this menu:\n')
                                if name != 'leave':
                                    sh = wb['client']
                                    r=2
                                    checkname = 'NO'
                                    while sh.cell(row=r, column=1).value != None:
                                        if sh.cell(row=r, column=1).value == name:
                                            checkname = 'YES'
                                            break
                                        r+=1
                                    if checkname == 'YES':
                                        w = xl.load_workbook('price_list.xlsx')
                                        sheet = w['price']
                                        r = 2
                                        print('Procedures with prices:')
                                        w = xl.load_workbook('price_list.xlsx')
                                        sp = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
                                        table = PrettyTable(sp)
                                        for a, b in sheet['A2': 'B6']:
                                            table.add_row(
                                                [a.value, b.value])
                                        print(table)
                                        print()
                                        while True:
                                            procname = input('Enter name of the procedure:\n').capitalize()
                                            r = 2
                                            checkproc = 'NO'
                                            while sheet.cell(row=r, column=1).value != None:
                                                if sheet.cell(row=r, column=1).value == procname:
                                                    checkproc = 'YES'
                                                    break
                                                r += 1
                                            if checkproc == 'YES':
                                                w2 = xl.load_workbook('client_info.xlsx')
                                                sh2 = w2[name]
                                                r = 2
                                                while sh2.cell(row=r, column=4).value != None:
                                                    r += 1
                                                w = xl.load_workbook('price_list.xlsx')
                                                sh = w['time']
                                                sp = [sh.cell(row=1, column=i).value for i in
                                                      range(1, sh.max_column + 1)]
                                                table = PrettyTable(sp)
                                                for a, b, c, d, e, f, g in sh['A2': 'G6']:
                                                    table.add_row(
                                                        [a.value, b.value, c.value, d.value, e.value, f.value, g.value])
                                                print(table)
                                                print()
                                                while True:
                                                    day = input('Choose the day:\n').capitalize()
                                                    days = ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday',
                                                            'Saturday')
                                                    if day in days:
                                                        ptime = input('Choose procedure time(example: 14:00):\n')
                                                        sh2.cell(row=r,
                                                                 column=4).value = procname + f', {day}, {ptime}'
                                                        print('Successfully bought!')
                                                        print()
                                                        w2.save('client_info.xlsx')
                                                        break
                                                    else:
                                                        print('Write correctly!')
                                                break
                                            else:
                                                print('Enter correct name of the procedure!')
                                                print()
                                        break
                                    else:
                                        print('Write correct name!')

                                else:
                                    print()
                                    break
                        elif mnum == 6:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('6. Looking for a procedure:\n')
                            while True:
                                wb = xl.load_workbook('price_list.xlsx')
                                sh = wb['price']
                                lfp = input("Which procedure are you looking for?\n").lower()
                                for i in range(2, sh.max_row + 1):
                                    if lfp.capitalize() == sh.cell(row=i, column=1).value:
                                        lfpi = i
                                        break
                                    else:
                                        lfpi = 'No'
                                if lfpi != 'No':
                                    sp = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
                                    table = PrettyTable(sp)
                                    table.add_row(
                                            [sh.cell(row=i, column=1).value, sh.cell(row=i, column=2).value])
                                    print(table)
                                    print()
                                    break
                                else:
                                    print('Write correct procedure name!')
                        elif mnum == 7:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print('7. Info about you:\n')
                            w = xl.load_workbook('sassistant_info.xlsx')
                            sh = w[log]
                            table = PrettyTable()
                            table.field_names = ['Info']
                            r=2
                            while sh.cell(row=r, column=1).value != None:
                                table.add_row([sh.cell(row=r, column=1).value])
                                r+=1
                            print(table)
                            print()
                            while True:
                                change = input('Do you want to change something?(Enter "yes" or "no")\n').lower()
                                if change == 'yes':
                                    while True:
                                        exchange = input('What exactly do you want to change? (Enter "name", "birthday" or "num")\n').lower()
                                        if exchange == 'name':
                                            newname = input('Enter your full name(example: Umurzakov Diyor):\n')
                                            sh.cell(row=2, column=1).value = newname
                                            break
                                        elif exchange == 'birthday':
                                            newbirth = input('Enter your birthday(example: 01.01.2000):\n')
                                            sh.cell(row=3, column=1).value = newbirth
                                            break
                                        elif exchange == 'num':
                                            newnum = int(input('Enter your phone number(without +996 and 0):\n'))
                                            sh.cell(row=4,column=1).value = newnum
                                            break
                                        else:
                                            print('Write correctly!\n')
                                    w.save('sassistant_info.xlsx')
                                    break
                                elif change == 'no':
                                    break
                                else:
                                    print('Write correctly!\n')

                        elif mnum == 8:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print("You left your account.")
                            main()
                            break
                        else:
                            os.system('cls' if os.name == 'nt' else 'clear')
                            print("Enter number from 1 to 7")
                    break
                else:
                    print('Write correct password\n')
            break
        else:
            print('Write correct login\n')


def main():
    while True:  # авторизация
        acc = input("If you have account, enter 'login', else enter 'create' or if you want to leave the program, enter 'leave':\n").lower()
        if acc == 'login':
            acctype = input("Enter type of account (Client, Manager, SAssistant, Director):\n").lower()
            if acctype == 'client':
                os.system('cls' if os.name == 'nt' else 'clear')
                client()
                break
            elif acctype == 'manager':
                os.system('cls' if os.name == 'nt' else 'clear')
                manager()
                break
            elif acctype == 'sassistant':
                os.system('cls' if os.name == 'nt' else 'clear')
                sassistant()
                break
            elif acctype == 'director':
                os.system('cls' if os.name == 'nt' else 'clear')
                director()
                break
            else:
                print("Sorry, we haven't found this type of account, please try again.\n")
        elif acc == 'create': # создание аккаунта
            acctype = input("Enter type of account (Client, Manager, SAssistant):\n").lower()
            if acctype == 'client':
                w = xl.load_workbook('loginpass.xlsx')
                newlog = input("Enter your name:\n")
                sh = w['client']
                mrow = sh.max_row + 1
                sh.cell(row=mrow, column=1).value = newlog
                newpass = input('Enter your password:\n')
                sh.cell(row=mrow, column=2).value = newpass
                w.save('loginpass.xlsx')
                os.system('cls' if os.name == 'nt' else 'clear')
                wb = xl.load_workbook('client_info.xlsx')
                wb.create_sheet(f'{newlog}')
                sh1=wb[f'{newlog}']
                sh1.cell(row=1, column=1).value='Info'
                sh1.cell(row=1, column=2).value = 'Timetable'
                sh1.cell(row=1, column=3).value = 'Visiting history'
                sh1.cell(row=1, column=4).value = 'Purchased procedures'
                sh1.cell(row=2, column=1).value = input('Enter your full name:\n')
                sh1.cell(row=3, column=1).value = 'Height: ' + input('Enter your height:\n')
                sh1.cell(row=4, column=1).value = 'Weight: ' + input('Enter your weight:\n')
                sh1.cell(row=5, column=1).value = 'Birthday: ' + input('Enter your birthday(example: 20.10.2000):\n')
                wb.save('client_info.xlsx')
                client()
                break
            elif acctype == 'manager':
                w = xl.load_workbook('loginpass.xlsx')
                newlog = input("Enter your name:\n")
                sh = w['manager']
                mrow = sh.max_row + 1
                sh.cell(row=mrow, column=1).value = newlog
                newpass = input('Enter your password:\n')
                sh.cell(row=mrow, column=2).value = newpass
                w.save('loginpass.xlsx')
                os.system('cls' if os.name == 'nt' else 'clear')
                manager()
                break
            elif acctype == 'sassistant':
                w = xl.load_workbook('loginpass.xlsx')
                newlog = input("Enter your name:\n")
                sh = w['sassistant']
                mrow = sh.max_row + 1
                sh.cell(row=mrow, column=1).value = newlog
                newpass = input('Enter your password:\n')
                sh.cell(row=mrow, column=2).value = newpass
                w.save('loginpass.xlsx')
                os.system('cls' if os.name == 'nt' else 'clear')
                sassistant()
                break
            else:
                os.system('cls' if os.name == 'nt' else 'clear')
                print("Sorry, we haven't found this type of account, please try again.\n")
        elif acc == 'leave':
            print()
            print('Good bye!')
            break
        else:
            os.system('cls' if os.name == 'nt' else 'clear')
            print('Please, write "login" or "create"')

main()